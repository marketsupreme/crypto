#!/usr/bin/env python
# coding: utf-8

from requests import get, patch
import json
import time as t
from datetime import date, datetime
import calendar
import pandas as pd
from openpyxl import load_workbook

def getCoinValue(coinDict) -> dict():

    #nested dictionary will store each coin's information
    portfolio = {k:{} for k in coinDict.keys()}

    #getting all crypto prices from Nomics and storing them in a JSON
    prices = get(f'https://api.nomics.com/v1/currencies/ticker?key=ea386addbac03f4bb67ceb1f333a8d0a&ids={",".join(coinDict)}&interval=1d&convert=USD&per-page=100&page=1')
    priceData = prices.json()

    

    #getting ethereum holdings and storing in a JSON
    # ethWallet = get('https://api.ethplorer.io/getAddressInfo/0xed8b4b3ba4fd5a175613859cab6ab8f010276a3a?apiKey=freekey')
    # ethData = ethWallet.json()
    # portfolio['ETH']['holdings']  = ethData['ETH']['balance']

    #adding the value to portfolio dictionary
    count = 0
    for key in coinDict.keys():
        portfolio[key]['price'] = round(float(priceData[count]['price']),2)
        portfolio[key]['holdings'] = coinDict[key]
        portfolio[key]['value'] = portfolio[key]['holdings']*portfolio[key]['price']
        count += 1
        
    return portfolio

def getStockValue(stockDict) -> dict():

    stockKey = '10SVU9Y1PENCIBZN'
    portfolio = {k:{} for k in stockDict.keys()}

    for ticker in stockDict.keys():
        prices = get(f'https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={ticker}&apikey={stockKey}')
        priceData = prices.json()
        portfolio[ticker]['price'] = round(float(priceData["Global Quote"]["05. price"]),2)
        portfolio[ticker]['holdings'] = stockDict[ticker]
        portfolio[ticker]['value'] = portfolio[ticker]['holdings']*portfolio[ticker]['price']
        t.sleep(15)
    
    return portfolio

def main():

    #stocks and crypto variables, available to be updated at anytime
    Coins = {'BTC': 1.06, 'ETH': 26.960005962, 'DOGE': 1809.826}
    Stocks = {'AMZN': 3,'ACB': 41,'ACNNF': 1000,'CNGGF': 200, 
            'CHALF': 217, 'EL': 5,'PLTR': 61,
            'PYPL': 6,'QS': 30,
            'SMG': 15,'TSLA': 3,
            'TLRY': 167, 'MRRCF': 700}

    #opening the readme to be displayed on Github pages
    file = open('README.md', 'w')

    #grabbing date information
    today = date.today().strftime("%m/%d/%y")
    time = datetime.now().strftime("%H:%M:%S")
    day = calendar.day_name[date.today().weekday()]

    #calling getValue function to find portfolio value
    coinDict = getCoinValue(Coins)
    stockDict = getStockValue(Stocks)

    #formating and rounding the total portfolio values
    coinValue = sum(coin['value'] for coin in coinDict.values() if coin)
    stockValue = sum(stock['value'] for stock in stockDict.values() if stock)
    totalValue = "{:,}".format(round(coinValue+stockValue,2))

    coinValue = "{:,}".format(round(coinValue,2))
    stockValue = "{:,}".format(round(stockValue,2))

    #writing to the readme with value and timestamp
    file.write(f'# Value: ${totalValue} as of {day}, {today} @ {time} \n\n')
    file.write(f'### Crypto Value: ${coinValue}\n\n')
    file.write(f'### Stock Value: ${stockValue}\n\n')

    file.write('#### Crypto Information \n')

    file.write('*Crypto prices* \n\n')
    #Coin prices
    for coin in Coins:
        file.write(f"{coin} Price = ${'{:,}'.format(coinDict[coin]['price'])},  \n")
    file.write('\n\n')
    
    file.write('*Crypto holdings* \n\n')
    #Coin holdings
    for coin in Coins:
        file.write(f"{coin} Holdings = {coinDict[coin]['holdings']}{coin},  \n")
    file.write('\n\n')

    file.write('#### Stock Information \n\n')

    file.write('*Stock prices* \n\n')
    #Stock prices
    for stock in Stocks:
        file.write(f"{stock} Price = ${'{:,}'.format(stockDict[stock]['price'])},  \n")
    file.write('\n\n')
    
    file.write('*Stock holdings* \n\n')
    #Stock holdings
    for stock in Stocks:
        file.write(f"{stock} Holdings = {stockDict[stock]['holdings']},  \n")
    file.write('\n\n')

    df = pd.DataFrame({'Date':[str(today)],'Value':[totalValue]})
    
    #writing the portfolio value to an excel file for data compilation
    wb = load_workbook(filename = 'value.xlsx')
    ws = wb['Sheet']
    newRowLoc = ws.max_row + 1
    ws.cell(column=1, row = newRowLoc, value =str(today))
    ws.cell(column=2, row = newRowLoc, value =str(totalValue))
    wb.save(filename='value.xlsx')
    wb.close()

    t.sleep(1)

main()
